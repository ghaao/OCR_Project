from difflib import SequenceMatcher
import openpyxl
from collections import OrderedDict
from itertools import repeat
from ocr_app.static.properties.enumeration import DTC, MTM
from ocr_app.spoiler.utils.stringutil import CharacterSeperator


class WordMapper():
    """
    1. Measures similarity by separating words used in receipts and recognized words into consonants and vowels
    ex) ㅈㅣㄴㄷㅏㄴㄹㅛ (진단료), ㅈㅣㄴㄷㅏㄴㅂㅣ (진단비) -> 0.71
    2. Compare the similarities of all the words in the list, select the MAX value, and map similarity plots above a certain value    
    """
    def __init__(self, item_words, table_contents, table_coords, dtc, tolerance_value=0.6, mtm = MTM.SINGLE_WORD):
        if MTM.NUMERIC == mtm:
            self._letter_number = [word for word in table_contents if len(word[0]) > 0]
            self._table_contents = [word[0] for word in table_contents if len(word[0]) > 0]
        elif MTM.SINGLE_WORD == mtm:
            self._table_contents = table_contents

        self._original_words = [word[1] for word in item_words]
        self._recognition_words = [word[0] for word in self._table_contents]
        self._tolerance_value = tolerance_value
        self._table_coords = table_coords        
        self._item_words = item_words
        self._dtc = dtc
        self._mtm = mtm
        self._filtered_result = []
        self._base_item_lines = []
        self._vertical_base_column = None
        self._vertical_base_column_width = None
        self._cost_items = []
        self._final_result = []

    def word_mapping(self):
        """
        return [float ratio, original_word, recognition_word]
        """
        results = []
        filtered_results = []
        character_seperator = CharacterSeperator()

        for original_word in self._original_words:
            character_seperator.set(original_word)
            character_seperator.run()
            source_word = character_seperator.get()
            ratio_list = []
            for recognition_word in self._recognition_words[:]:
                character_seperator.set(recognition_word)
                character_seperator.run()
                target_word = character_seperator.get()
                ratio = SequenceMatcher(None, source_word, target_word).ratio()
                ratio_list.append([ratio, original_word, recognition_word])

                if ratio == 1:
                    self._recognition_words.remove(recognition_word)
                    break

            max_ratio_word = max(ratio_list)
            results.append(max_ratio_word)

        for result in results:
            if result[0] >= self._tolerance_value:
                self._filtered_result.append(result)

        # 중복 제거
        dups = [item[2] for item in self._filtered_result]
        w_count = {}
        for lst in dups:
            try: w_count[lst] += 1
            except: w_count[lst] = 1

        duplication = [temp for temp in w_count.items() if temp[1] > 1]

        bullets = []
        for mapping in self._filtered_result:
            for target in duplication:
                if mapping[2] == target[0]:
                    bullets.append(mapping)
                    self._filtered_result.remove(mapping)
        # filtered_list에서 중복된 건은 모두 지우고 max값으로 추가, 이외 건은 다시 mapping        
        bullet_words = list(set(temp[2] for temp in bullets))

        for word in bullet_words[:]:
            max_word = max(bull for bull in bullets if bull[2] == word)
            self._filtered_result.append(max_word)            
            bullets.remove(max_word)

            try:
                self._recognition_words.remove(max_word[2])      
            except ValueError:                
                pass

        #  End of recursive method
        if len(bullets) == 0 :
            # recursive end
            # print("###", DTC(self._dtc).name, "mapping")
            pass            
        else:
            self._original_words = [bull[1] for bull in bullets]
            self.word_mapping()

    def _rbm_item_fine_tune(self):
        """
        Reimbursement receipt item finetune.
        """
        v_coord = []
        h_coord = []
        for mapp in self._filtered_result:
            for content in self._table_contents:
                if mapp[2] == content[0]:
                    mapp.extend(content[1:])
                    for word in self._item_words:
                        if mapp[1]==word[1]:
                            if word[0] == "v_menu":
                                if mapp[0] == 1.0:                            
                                    v_coord.append(mapp)
                            else:
                                if mapp[0] == 1.0:                                                    
                                    h_coord.append(mapp)
        # 인식 문자중에 가장 왼쪽에 있는 컬럼을 기준으로 세로 메뉴를 추출 (가장 왼쪽에 있는 칸이 가장 넓음.)

        self._vertical_base_column = min(self._filtered_result, key=lambda item: item[3])
        print("기본항목:", self._vertical_base_column)

        self._base_item_lines = []
        for table in self._table_coords:    
            if self._vertical_base_column[3] - 20 < table[0] and self._vertical_base_column[5] + 20 > table[2]:
                flg = False
                for mapp in self._filtered_result:            
                    if mapp[3:] == table:                        
                        self._base_item_lines.append(mapp)
                        flg = True
                        break
                if flg is not True:                
                    self._base_item_lines.append([0,"",""]+table)
                
        self._base_item_lines = [i for i in sorted(self._base_item_lines, key=lambda line: line[4], reverse = False)  if i[5]-i[3] > i[6]-i[4] ]

        for base_item in self._base_item_lines[:]:    
            if base_item[0] > 0:
                break   
            self._base_item_lines.remove(base_item)

        for base_item in self._base_item_lines[::-1]:
            if base_item[2] == "합계":
                self._base_item_lines.remove(base_item)
                break
            if base_item[0] > 0:               
                break
            self._base_item_lines.remove(base_item)

        # 투약및조제료 빈칸 채우기
        self._vertical_base_column_width = (self._vertical_base_column[5]-self._vertical_base_column[3])/2

        i = 0
        j = 0
        # 세로 순서대로 정렬하여 순서로 파악 가능
        for base_item in self._base_item_lines:    
            if self._vertical_base_column_width > base_item[5]-base_item[3]:        
                # 인식글자가 없고
                if base_item[1] == '':
                    # 첫번째 투약및조제료 칸이면
                    if i < 3 :
                        # 기준필드의 중심점이 포함이면 투약및조제료
                        if base_item[3] < self._vertical_base_column[3] + self._vertical_base_column_width and self._vertical_base_column[3] + self._vertical_base_column_width < base_item[5]:
                            base_item[1] = '투약및조제료'
                        elif i < 2: 
                            base_item[1] = '행위료'
                        else:
                            base_item[1] = '약품비'
                    # 두번째 주사료 칸이면
                    else:
                        if base_item[3] < self._vertical_base_column[3] + self._vertical_base_column_width and self._vertical_base_column[3] + self._vertical_base_column_width < base_item[5]:
                            base_item[1] = '주사료'
                        elif i < 5: 
                            base_item[1] = '행위료'
                        else:
                            base_item[1] = '약품비'
                i = i+1
            # 진단료 순서대료 CT, MRI, PET
            if base_item[1] == '진단료':
                if j==0:
                    base_item[1] = 'CT진단료'
                elif j==1:
                    base_item[1] = 'MRI진단료'
                elif j==2:
                    base_item[1] = 'PET진단료'
                j = j +1


        for base_item in self._base_item_lines[:]:    
            if self._vertical_base_column_width > base_item[5]-base_item[3]:  
                if base_item[1] == "투약및조제료": 
                    for base_item2 in self._base_item_lines:
                        if self._vertical_base_column_width > base_item2[5]-base_item2[3]: 
                            if base_item[4] - 20 < base_item2[4] and base_item[6] + 20 > base_item2[6] and base_item2[1] != "투약및조제료":                        
                                base_item2[1] = "투약및조제료_" + base_item2[1]
                                                        
                if base_item[1] == "주사료": 
                    for base_item2 in self._base_item_lines:
                        if self._vertical_base_column_width > base_item2[5]-base_item2[3]: 
                            if base_item[4] - 20 < base_item2[4] and base_item[6] + 20 > base_item2[6] and base_item2[1] != "주사료":                        
                                base_item2[1] = "주사료_" + base_item2[1] 

        for base_item in self._base_item_lines[:]:
            if base_item[1] == "투약및조제료" or  base_item[1] == "주사료":
                self._base_item_lines.remove(base_item)
        # 추가로 투약 및 조제로 주사료가 없는 경우 로직 만들어야 함
        
            


    def _rbm_cost_fine_tune(self):
    
        self._cost_items = []
        _column_width = self._vertical_base_column[5] - self._vertical_base_column[3]        
        _center_point = self._vertical_base_column[3] + _column_width/2 + _column_width

        for i in range(5):
            first_box_flg = True
            for content in self._table_contents[:]:                
                # vertical zone limit
                if self._base_item_lines[0][4] - 20 < content[2] and self._base_item_lines[-1][4] + 20 > content[2]  :                   
                    # Horizontal zone limit
                    if content[1] < _center_point and content[3] > _center_point:                        
                        if isinstance(content[0],int):                            
                            content[0] = str(content[0])
                        _digit = int(( lambda x : '0' if '' == x else x )(''.join(filter(lambda x: x.isdigit(), content[0]))))
                        content[0] = _digit
                        self._cost_items.append([i, content])
                        # 처음 나온 박스 값을 기준값으로 고정
                        if first_box_flg:
                            _column_width = content[3] - content[1]       
                            _center_point = content[1] + _column_width/2 
                            first_box_flg =False
                        
                
            _center_point += _column_width


    def _item_cost_mapping(self):
        
        code_mapper = lambda x : '본인부담금' if x==0 else ('공단부담금' if x==1 else ('전액본인부담' if x==2 else ('선택진료료' if x==3 else ('선택진료료이외' if x==4 else '오류발생'))))
        for cost_item in self._cost_items:
            exist_flg = False
            center = (cost_item[1][4] - cost_item[1][2]) / 2 + cost_item[1][2]
            for i, base_item in enumerate(self._base_item_lines):
                if base_item[6] > center and base_item[4] < center:
                    self._final_result.append([(lambda x : '인식실패'+str(i) if x == '' else x)(base_item[1]), code_mapper(cost_item[0]), cost_item[1][0]])
                    exist_flg = True
                    break
            if exist_flg is not True:
                self._final_result.append(['인식실패'+str(i), code_mapper(cost_item[0]), cost_item[1][0]])

    def _certi_result_mapping(self):        
        for filtered in self._filtered_result:
            filtered.extend(["", ""])
            for item in self._item_words:                
                if filtered[1] == item[1]:                    
                    filtered[3] = item[0] 
            # 값 셋팅
            if self._mtm == MTM.NUMERIC:                           
                for contents in self._letter_number:                
                    if contents[0][0] == filtered[2]:
                        if len(contents[1]) > 0:
                            filtered[4] = contents[1][0]
            elif self._mtm == MTM.SINGLE_WORD: 
                    pass 
        filtered_result = self._filtered_result.copy()
            # 중복 중 정확도가 가장 큰 값 
        for word in filtered_result:
            for source in filtered_result:
                if word[3] == source[3]:
                    if word[0] > source[0]:
                        filtered_result.remove(source)
                    elif word[0] < source[0]:
                        filtered_result.remove(word)
        return filtered_result
                

    def get_mapping_results(self):        
        if self._dtc == DTC.MEDICAL_CERTIFICATE:
            filtered_result = self._certi_result_mapping()
        elif self._dtc == DTC.MEDICAL_RECEIPT:
            self._rbm_item_fine_tune()
            self._rbm_cost_fine_tune()
            self._item_cost_mapping()
            filtered_result = self._final_result.copy()
        
        return filtered_result

    def get_mapping_results_excel(self):
        wb = openpyxl.Workbook()
        sheet = wb.active

        # _vertical = list(set([i[0] for i in self._final_result]))
        _vertical = list(OrderedDict(zip([i[0] for i in self._final_result], repeat(None))))

        for i, result in enumerate(_vertical):    
            sheet['A'+str(i+2)] = result
            for cost_vlu in self._final_result:
                if cost_vlu[1] == '본인부담금' and result == cost_vlu[0]:
                    sheet.cell(row=i+2, column=2).value = cost_vlu[2]
                elif cost_vlu[1] == '공단부담금' and result == cost_vlu[0]:
                    sheet.cell(row=i+2, column=3).value = cost_vlu[2]
                elif cost_vlu[1] == '전액본인부담' and result == cost_vlu[0]:
                    sheet.cell(row=i+2, column=4).value = cost_vlu[2]
                elif cost_vlu[1] == '선택진료료' and result == cost_vlu[0]:
                    sheet.cell(row=i+2, column=5).value = cost_vlu[2]
                elif cost_vlu[1] == '선택진료료이외' and result == cost_vlu[0]:
                    sheet.cell(row=i+2, column=6).value = cost_vlu[2]

        sheet['B1'] = '본인부담금'
        sheet['C1'] = '공단부담금'
        sheet['D1'] = '전액본인부담'
        sheet['E1'] = '선택진료료'
        sheet['F1'] = '선택진료료이외'
        wb.save('result.xlsx')   

